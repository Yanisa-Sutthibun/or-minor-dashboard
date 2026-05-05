"""
Minor OR Export — สร้างไฟล์ Excel สรุปสถิติพร้อมกราฟ
แยกเป็นไฟล์ต่างหากเพื่อไม่ให้ minor_or_db.py ใหญ่เกินไป
"""
import io
import pandas as pd


def export_summary_excel(get_summary_fn, export_cases_fn, div_name_fn,
                         date_from=None, date_to=None) -> bytes:
    """สร้างไฟล์ Excel สรุปสถิติพร้อมกราฟ — return bytes พร้อม download.

    Parameters:
        get_summary_fn: function get_summary(date_from, date_to)
        export_cases_fn: function export_cases_csv(date_from, date_to)
        div_name_fn: function div_name(code) → Thai name
        date_from, date_to: date range strings
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = export_cases_fn(date_from, date_to)
    summary = get_summary_fn(date_from, date_to)
    wb = Workbook()

    # --- Styles ---
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2E7D32')
    data_font = Font(name='Arial', size=10)
    num_fmt = '#,##0'
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    title_font = Font(name='Arial', bold=True, size=14, color='1B5E20')

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    def style_data(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = data_font
            cell.border = border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # ========== Sheet 1: สรุปภาพรวม ==========
    ws1 = wb.active
    ws1.title = 'สรุปภาพรวม'

    period_text = ''
    if date_from and date_to:
        period_text = f'{date_from} ถึง {date_to}'
    elif date_from:
        period_text = f'ตั้งแต่ {date_from}'
    else:
        period_text = 'ทั้งหมด'

    ws1['A1'] = 'สรุปสถิติห้องผ่าตัดเล็ก (Minor OR)'
    ws1['A1'].font = title_font
    ws1['A2'] = f'ช่วงเวลา: {period_text}'
    ws1['A2'].font = Font(name='Arial', size=11, color='666666')

    # KPI table
    kpi_data = [
        ['ตัวชี้วัด', 'จำนวน'],
        ['เคสทั้งหมด', summary['total']],
        ['ผ่าเสร็จ', summary['completed']],
        ['ยกเลิก', summary['cancelled']],
        ['OPD', summary.get('n_opd', 0)],
        ['IPD', summary.get('n_ipd', 0)],
        ['เคสนัดหมาย', summary.get('n_set', 0)],
        ['Walk-in', summary.get('n_walkin', 0)],
        ['ค่าหัตถการรวม (บาท)', summary.get('total_treatment', 0)],
        ['รายได้รวม (บาท)', summary.get('total_revenue', 0)],
        ['ส่งชิ้นเนื้อ (ราย)', summary.get('n_patho_sent', 0)],
        ['ค่าชิ้นเนื้อรวม (บาท)', summary.get('total_patho', 0)],
    ]
    for r, row_data in enumerate(kpi_data, start=4):
        for c, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=r, column=c, value=val)
            if r == 4:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = data_font
                if c == 2 and isinstance(val, (int, float)):
                    cell.number_format = num_fmt
            cell.border = border

    # AI Accuracy
    ai_data = summary.get('ai_accuracy', {})
    if ai_data and ai_data.get('count', 0) > 0:
        r_start = 18
        ws1.cell(row=r_start, column=1, value='AI Prediction Accuracy').font = Font(
            name='Arial', bold=True, size=12, color='1565C0')
        ai_rows = [
            ['ตัวชี้วัด', 'ค่า'],
            ['จำนวนเคสที่วัดได้', ai_data.get('count', 0)],
            ['MAE (นาที)', round(ai_data.get('mae', 0), 1)],
            ['MAPE (%)', round(ai_data.get('mape', 0), 1)],
            ['ภายใน +/-15 นาที (%)', round(ai_data.get('within_15', 0), 1)],
            ['ภายใน +/-30 นาที (%)', round(ai_data.get('within_30', 0), 1)],
        ]
        for r, row_data in enumerate(ai_rows, start=r_start + 1):
            for c, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=r, column=c, value=val)
                if r == r_start + 1:
                    cell.font = header_font
                    cell.fill = PatternFill('solid', fgColor='1565C0')
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.font = data_font
                cell.border = border

    auto_width(ws1)

    # ========== Sheet 2: Top Operations + กราฟ ==========
    if not df.empty:
        ws2 = wb.create_sheet('หัตถการยอดนิยม')
        in_hours = df[df['patient_type'] != 'นอกเวลา']
        top_ops = in_hours.groupby('procedure_name').size().reset_index(name='จำนวน')
        top_ops = top_ops.sort_values('จำนวน', ascending=False).head(10)

        ws2['A1'] = 'หัตถการที่ทำบ่อยที่สุด (Top 10)'
        ws2['A1'].font = title_font

        headers = ['หัตถการ', 'จำนวน (เคส)']
        for c, h in enumerate(headers, start=1):
            ws2.cell(row=3, column=c, value=h)
        style_header(ws2, 3, 2)

        for r, (_, row) in enumerate(top_ops.iterrows(), start=4):
            ws2.cell(row=r, column=1, value=row['procedure_name']).font = data_font
            ws2.cell(row=r, column=2, value=int(row['จำนวน'])).font = data_font
            ws2.cell(row=r, column=2).number_format = num_fmt
            style_data(ws2, r, 2)

        if len(top_ops) > 0:
            chart = BarChart()
            chart.type = 'col'
            chart.title = 'Top 10 หัตถการ'
            chart.y_axis.title = 'จำนวน (เคส)'
            chart.style = 10
            chart.width = 22
            chart.height = 14
            data_ref = Reference(ws2, min_col=2, min_row=3, max_row=3 + len(top_ops))
            cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(top_ops))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            ws2.add_chart(chart, 'D3')

        auto_width(ws2)

    # ========== Sheet 3: แยกตามสาขา + กราฟ ==========
    if not df.empty:
        ws3 = wb.create_sheet('แยกตามสาขา')
        div_stats = in_hours.copy()
        if 'division_name' not in div_stats.columns and 'division_code' in div_stats.columns:
            div_stats['division_name'] = div_stats['division_code'].apply(div_name_fn)
        div_grp = div_stats.groupby('division_name').size().reset_index(name='จำนวน')
        div_grp = div_grp.sort_values('จำนวน', ascending=False)

        ws3['A1'] = 'จำนวนเคสแยกตามสาขา'
        ws3['A1'].font = title_font

        for c, h in enumerate(['สาขา', 'จำนวน (เคส)'], start=1):
            ws3.cell(row=3, column=c, value=h)
        style_header(ws3, 3, 2)

        for r, (_, row) in enumerate(div_grp.iterrows(), start=4):
            ws3.cell(row=r, column=1, value=row['division_name']).font = data_font
            ws3.cell(row=r, column=2, value=int(row['จำนวน'])).font = data_font
            style_data(ws3, r, 2)

        if len(div_grp) > 0:
            pie = PieChart()
            pie.title = 'สัดส่วนเคสแยกตามสาขา'
            pie.style = 10
            pie.width = 18
            pie.height = 14
            data_ref = Reference(ws3, min_col=2, min_row=3, max_row=3 + len(div_grp))
            cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=3 + len(div_grp))
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            ws3.add_chart(pie, 'D3')

        auto_width(ws3)

    # ========== Sheet 4: ประเภทผู้ป่วย + กราฟ ==========
    if not df.empty:
        ws4 = wb.create_sheet('ประเภทผู้ป่วย')
        pt_grp = df.groupby('patient_type').size().reset_index(name='จำนวน')
        pt_grp = pt_grp.sort_values('จำนวน', ascending=False)

        ws4['A1'] = 'จำนวนเคสแยกตามประเภทผู้ป่วย'
        ws4['A1'].font = title_font

        for c, h in enumerate(['ประเภท', 'จำนวน (เคส)'], start=1):
            ws4.cell(row=3, column=c, value=h)
        style_header(ws4, 3, 2)

        for r, (_, row) in enumerate(pt_grp.iterrows(), start=4):
            ws4.cell(row=r, column=1, value=row['patient_type'] or 'ไม่ระบุ').font = data_font
            ws4.cell(row=r, column=2, value=int(row['จำนวน'])).font = data_font
            style_data(ws4, r, 2)

        if len(pt_grp) > 0:
            chart = BarChart()
            chart.type = 'col'
            chart.title = 'ประเภทผู้ป่วย'
            chart.style = 10
            chart.width = 18
            chart.height = 12
            data_ref = Reference(ws4, min_col=2, min_row=3, max_row=3 + len(pt_grp))
            cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=3 + len(pt_grp))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws4.add_chart(chart, 'D3')

        auto_width(ws4)

    # ========== Sheet 5: รายได้แยก operation ==========
    if not df.empty:
        ws5 = wb.create_sheet('รายได้แยก operation')
        rev = in_hours[in_hours['status'] == 'discharged'].copy()
        if not rev.empty and 'treatment_cost' in rev.columns:
            rev['treatment_cost'] = pd.to_numeric(
                rev.get('treatment_cost', 0), errors='coerce').fillna(0).astype(int)
            rev_grp = rev.groupby('procedure_name').agg(
                count=('procedure_name', 'size'),
                total=('treatment_cost', 'sum'),
                avg=('treatment_cost', 'mean'),
            ).reset_index().sort_values('total', ascending=False)

            ws5['A1'] = 'รายได้ค่าหัตถการแยกตาม Operation'
            ws5['A1'].font = title_font

            headers = ['หัตถการ', 'จำนวน', 'รายได้รวม (บาท)', 'เฉลี่ย (บาท)']
            for c, h in enumerate(headers, start=1):
                ws5.cell(row=3, column=c, value=h)
            style_header(ws5, 3, 4)

            for r, (_, row) in enumerate(rev_grp.iterrows(), start=4):
                ws5.cell(row=r, column=1, value=row['procedure_name']).font = data_font
                ws5.cell(row=r, column=2, value=int(row['count'])).font = data_font
                ws5.cell(row=r, column=3, value=int(row['total'])).font = data_font
                ws5.cell(row=r, column=3).number_format = num_fmt
                ws5.cell(row=r, column=4, value=int(row['avg'])).font = data_font
                ws5.cell(row=r, column=4).number_format = num_fmt
                style_data(ws5, r, 4)

            auto_width(ws5)

    # ========== Sheet 6: ข้อมูลดิบ ==========
    if not df.empty:
        ws6 = wb.create_sheet('ข้อมูลดิบ')
        ws6['A1'] = 'ข้อมูลเคสทั้งหมด (Raw Data)'
        ws6['A1'].font = title_font

        col_map = {
            'op_date': 'วันที่', 'name': 'ชื่อผู้ป่วย', 'hn': 'HN', 'an': 'AN',
            'procedure_name': 'หัตถการ', 'surgeon_name': 'แพทย์',
            'division_name': 'สาขา', 'patient_type': 'ประเภท',
            'case_category': 'หมวด', 'status': 'สถานะ',
            'room_no': 'ห้อง', 'scrub_nurse': 'Scrub', 'circ_nurse': 'Circulate',
            'ai_predicted_min': 'AI ทำนาย (นาที)',
            'actual_duration_min': 'เวลาจริง (นาที)',
            'arrived_at': 'มาถึง', 'in_or_at': 'เข้าห้อง',
            'op_end_at': 'เสร็จ', 'discharged_at': 'จำหน่าย',
        }
        cols_to_use = [c for c in col_map if c in df.columns]
        df_out = df[cols_to_use].rename(columns=col_map)

        for c, col_name in enumerate(df_out.columns, start=1):
            ws6.cell(row=3, column=c, value=col_name)
        style_header(ws6, 3, len(df_out.columns))

        for r, (_, row) in enumerate(df_out.iterrows(), start=4):
            for c, val in enumerate(row, start=1):
                cell = ws6.cell(row=r, column=c, value=val)
                cell.font = data_font
                cell.border = border

        auto_width(ws6)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
